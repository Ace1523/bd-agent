#!/usr/bin/env python3
"""Import ATL_BD enriched Excel outputs into docs/connections.json for the LinkedIn Connections dashboard."""

import json
import os
from collections import defaultdict
from datetime import date

import openpyxl

import re

ATL_BD_DIR = os.path.expanduser("~/Desktop/ATL_BD")
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
REGIONS = {
    "ATL": {
        "display_name": "Atlanta",
        "file": os.path.join(ATL_BD_DIR, "regions", "ATL", "ATL Region Targeting - Enriched.xlsx"),
    },
    "DC": {
        "display_name": "Washington DC",
        "file": os.path.join(ATL_BD_DIR, "regions", "DC", "DC Region Targeting - Enriched.xlsx"),
    },
    "UK": {
        "display_name": "United Kingdom",
        "file": os.path.join(DATA_DIR, "charlie-antelme-connections.xlsx"),
        "parser": "charlie",
    },
}
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "docs", "connections.json")

# Senior leadership role pattern — used to split contacts into key leaders vs other connections
SENIOR_REGEX = re.compile(
    r"(chief|ceo|cfo|coo|cto|cio|ciso|cmo|chro|president|"
    r"vice\s*president|\bvp\b|svp|evp|director|general\s*manager|"
    r"managing\s*director|head\s+of|group\s+head|senior\s+director|"
    r"co-?founder|founder|partner|counsel|sector\s+lead|"
    r"regional\s+director|country\s+director|executive\s+director)",
    re.IGNORECASE,
)
DASHBOARD_PATH = os.path.join(os.path.dirname(__file__), "data", "dashboard.json")

# Companies to exclude entirely (will never be prospects)
EXCLUDED_COMPANIES = [
    "booz allen hamilton",
]


def load_overwatch_prospects():
    """Load existing Overwatch prospect names for pipeline cross-referencing."""
    if not os.path.exists(DASHBOARD_PATH):
        return set()
    with open(DASHBOARD_PATH) as f:
        data = json.load(f)
    names = set()
    for run in data.get("discovery_runs", []):
        for p in run.get("prospects", []):
            name = p.get("company_name", "") or p.get("name", "")
            if name:
                names.add(name.strip().lower())
    return names


def fuzzy_match(company_name, prospects):
    """Check if a company name matches any Overwatch prospect (case-insensitive, partial)."""
    lower = company_name.strip().lower()
    for p in prospects:
        if lower == p or lower in p or p in lower:
            return p
    return None


def parse_region(region_key, config, prospects):
    """Parse one enriched Excel workbook into a region dict."""
    filepath = config["file"]
    if not os.path.exists(filepath):
        print(f"  WARNING: {filepath} not found, skipping {region_key}")
        return None

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    # --- Enriched Targets ---
    ws = wb["Enriched Targets"]
    companies = {}  # company_name -> company dict
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        company_name = (row[0] or "").strip().replace("*", "")
        if not company_name:
            continue
        if company_name.lower() in EXCLUDED_COMPANIES:
            continue
        if company_name not in companies:
            companies[company_name] = {
                "company_name": company_name,
                "industry": (row[6] or "").strip(),
                "location": (row[7] or "").strip(),
                "revenue_millions": row[4] if isinstance(row[4], (int, float)) else None,
                "leaders": [],
                "company_connections": [],
                "military_affiliations": [],
                "has_connections": False,
                "connection_count": 0,
                "partner_names": [],
                "pipeline_match": None,
            }
        leader_name = (row[1] or "").strip().replace("*", "")
        role = (row[2] or "").strip().replace("*", "")
        mg_poc = (row[3] or "").strip().replace("*", "") if row[3] else None
        if leader_name:
            companies[company_name]["leaders"].append({
                "name": leader_name,
                "role": role,
                "mg_point_of_contact": mg_poc,
                "is_military": False,  # updated below
            })

    # --- Company Connections ---
    ws2 = wb["Company Connections"]
    partner_counts = defaultdict(set)  # partner -> set of company names
    for i, row in enumerate(ws2.iter_rows(values_only=True)):
        if i == 0:
            continue
        target = (row[0] or "").strip()
        conn_name = (row[1] or "").strip()
        position = (row[2] or "").strip()
        partner = (row[4] or "").strip()
        if target in companies:
            companies[target]["company_connections"].append({
                "connection_name": conn_name,
                "position": position,
                "mg_partner": partner,
            })
            companies[target]["has_connections"] = True
            if partner:
                partner_counts[partner].add(target)

    # Also track direct leader matches from column D (may be semicolon-delimited)
    for comp in companies.values():
        for leader in comp["leaders"]:
            poc_raw = leader["mg_point_of_contact"]
            if poc_raw:
                comp["has_connections"] = True
                # Split semicolon-delimited partner names
                pocs = [p.strip() for p in poc_raw.split(";") if p.strip()]
                leader["mg_point_of_contact"] = pocs[0] if len(pocs) == 1 else poc_raw
                for p in pocs:
                    partner_counts[p].add(comp["company_name"])

    # Propagate MG contacts across companies for multi-board leaders
    # If a leader has an MG POC at one company, apply it everywhere they appear
    leader_pocs = {}  # leader_name -> mg_point_of_contact
    for comp in companies.values():
        for leader in comp["leaders"]:
            if leader["mg_point_of_contact"]:
                existing = leader_pocs.get(leader["name"])
                if existing:
                    # Merge POCs
                    all_pocs = set(p.strip() for p in existing.split(";") if p.strip())
                    all_pocs.update(p.strip() for p in leader["mg_point_of_contact"].split(";") if p.strip())
                    leader_pocs[leader["name"]] = "; ".join(sorted(all_pocs))
                else:
                    leader_pocs[leader["name"]] = leader["mg_point_of_contact"]
    # Apply propagated POCs
    propagated = 0
    for comp in companies.values():
        for leader in comp["leaders"]:
            if not leader["mg_point_of_contact"] and leader["name"] in leader_pocs:
                leader["mg_point_of_contact"] = leader_pocs[leader["name"]]
                comp["has_connections"] = True
                for p in leader["mg_point_of_contact"].split(";"):
                    p = p.strip()
                    if p:
                        partner_counts[p].add(comp["company_name"])
                propagated += 1
    if propagated:
        print(f"  Propagated MG contacts to {propagated} multi-board leader entries")

    # Update connection counts and partner names per company
    for comp in companies.values():
        conns = comp["company_connections"]
        comp["connection_count"] = len(conns)
        # Merge partner names from both connections and direct leader matches
        partner_set = set(c["mg_partner"] for c in conns if c["mg_partner"])
        for leader in comp["leaders"]:
            if leader["mg_point_of_contact"]:
                for p in leader["mg_point_of_contact"].split(";"):
                    p = p.strip()
                    if p:
                        partner_set.add(p)
        comp["partner_names"] = sorted(partner_set)

    # --- Military Affiliations ---
    ws3 = wb["Military Affiliations"]
    for i, row in enumerate(ws3.iter_rows(values_only=True)):
        if i == 0:
            continue
        company_name = (row[0] or "").strip()
        leader_name = (row[1] or "").strip()
        if company_name in companies:
            companies[company_name]["military_affiliations"].append({
                "name": leader_name,
                "rank": (row[3] or "").strip(),
                "branch": (row[4] or "").strip(),
                "details": (row[5] or "").strip(),
            })
            # Mark leaders as military
            for leader in companies[company_name]["leaders"]:
                if leader["name"] == leader_name:
                    leader["is_military"] = True

    # --- Pipeline matching ---
    for comp in companies.values():
        match = fuzzy_match(comp["company_name"], prospects)
        if match:
            # Find original-cased name from prospects
            comp["pipeline_match"] = comp["company_name"]

    wb.close()

    # --- Filter: keep companies with MG connections OR military affiliations ---
    all_companies_count = len(companies)
    connected_companies = {k: v for k, v in companies.items()
                           if v["has_connections"] or v["military_affiliations"]}

    # --- Summary stats ---
    total_companies = all_companies_count
    connected = len(connected_companies)
    total_connections = sum(c["connection_count"] for c in connected_companies.values())
    coverage_pct = round((connected / total_companies * 100), 1) if total_companies else 0

    # Partner breakdown
    partner_breakdown = sorted(
        [{"partner": p, "company_count": len(comps)} for p, comps in partner_counts.items()],
        key=lambda x: x["company_count"],
        reverse=True,
    )

    # Sort by connection count desc, then alphabetically
    company_list = sorted(
        connected_companies.values(),
        key=lambda c: (-c["connection_count"], c["company_name"]),
    )

    return {
        "name": region_key,
        "display_name": config["display_name"],
        "total_companies": total_companies,
        "total_connections": total_connections,
        "connected_companies": connected,
        "coverage_pct": coverage_pct,
        "companies": company_list,
        "partner_breakdown": partner_breakdown,
    }


def parse_charlie_region(region_key, config, prospects):
    """Parse Charlie Antelme's single-sheet LinkedIn export into a region dict.

    Splits contacts into key leaders (senior roles) and other connections (non-senior).
    Columns: 0=Company, 1=Priority, 2=Name, 3=Role, 4=MG Contact, 5=Focus,
             6=Revenue, 7=Rev Year, 8=Industry, 9=Location.
    """
    filepath = config["file"]
    if not os.path.exists(filepath):
        print(f"  WARNING: {filepath} not found, skipping {region_key}")
        return None

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    companies = {}
    partner_counts = defaultdict(set)
    leader_count = 0
    connection_count = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        company_name = (row[0] or "").strip().replace("*", "")
        name = (row[2] or "").strip().replace("*", "")
        role = (row[3] or "").strip().replace("*", "")
        mg_contact = (row[4] or "").strip().replace("*", "")
        revenue = row[6] if len(row) > 6 and isinstance(row[6], (int, float)) else None
        industry = (row[8] or "").strip() if len(row) > 8 else ""
        location = (row[9] or "").strip() if len(row) > 9 else ""

        if not company_name or not name:
            continue
        if company_name.lower() in EXCLUDED_COMPANIES:
            continue

        if company_name not in companies:
            companies[company_name] = {
                "company_name": company_name,
                "industry": industry,
                "location": location,
                "revenue_millions": revenue,
                "leaders": [],
                "company_connections": [],
                "military_affiliations": [],
                "has_connections": True,
                "connection_count": 0,
                "partner_names": [],
                "pipeline_match": None,
            }

        if SENIOR_REGEX.search(role):
            # Key leader
            companies[company_name]["leaders"].append({
                "name": name,
                "role": role,
                "mg_point_of_contact": mg_contact,
                "is_military": False,
            })
            leader_count += 1
        else:
            # Other connection
            companies[company_name]["company_connections"].append({
                "connection_name": name,
                "position": role,
                "mg_partner": mg_contact,
            })
            companies[company_name]["connection_count"] += 1
            connection_count += 1

        if mg_contact:
            partner_counts[mg_contact].add(company_name)

    wb.close()

    # Update partner names per company
    for comp in companies.values():
        partner_set = set()
        for leader in comp["leaders"]:
            if leader["mg_point_of_contact"]:
                for p in leader["mg_point_of_contact"].split(";"):
                    p = p.strip()
                    if p:
                        partner_set.add(p)
        for conn in comp["company_connections"]:
            if conn["mg_partner"]:
                partner_set.add(conn["mg_partner"])
        comp["partner_names"] = sorted(partner_set)

    # Pipeline matching
    for comp in companies.values():
        match = fuzzy_match(comp["company_name"], prospects)
        if match:
            comp["pipeline_match"] = comp["company_name"]

    # Summary stats
    total_companies = len(companies)
    total_connections = sum(c["connection_count"] for c in companies.values())
    partner_breakdown = sorted(
        [{"partner": p, "company_count": len(comps)} for p, comps in partner_counts.items()],
        key=lambda x: x["company_count"],
        reverse=True,
    )
    company_list = sorted(companies.values(), key=lambda c: (-c["connection_count"], c["company_name"]))

    print(f"  {leader_count} key leaders, {connection_count} other connections across {total_companies} companies")

    return {
        "name": region_key,
        "display_name": config["display_name"],
        "total_companies": total_companies,
        "total_connections": total_connections,
        "connected_companies": total_companies,
        "coverage_pct": 100.0,
        "companies": company_list,
        "partner_breakdown": partner_breakdown,
    }


def main():
    print("Loading Overwatch prospects for pipeline matching...")
    prospects = load_overwatch_prospects()
    print(f"  Found {len(prospects)} prospects")

    regions = []
    for key, config in REGIONS.items():
        print(f"Parsing {key} ({config['display_name']})...")
        parser = config.get("parser", "enriched")
        if parser == "charlie":
            region = parse_charlie_region(key, config, prospects)
        else:
            region = parse_region(key, config, prospects)
        if region:
            regions.append(region)
            print(f"  {region['total_companies']} companies, {region['connected_companies']} connected, {region['coverage_pct']}% coverage")

    # Collect all unique partners
    all_partners = set()
    for r in regions:
        for pb in r["partner_breakdown"]:
            all_partners.add(pb["partner"])

    output = {
        "last_updated": date.today().isoformat(),
        "regions": regions,
        "partner_count": len(all_partners),
    }

    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\nWrote {OUTPUT_PATH}")
    total_cos = sum(r["total_companies"] for r in regions)
    total_conns = sum(r["total_connections"] for r in regions)
    print(f"  Total: {total_cos} companies, {total_conns} connections, {len(all_partners)} partners, {len(regions)} regions")


if __name__ == "__main__":
    main()
